#!/usr/bin/env python3
"""生成两份排放因子数据文件：

    docs/data/scoring/fuel_emission_factors.json         燃料燃烧因子（IPCC 2006 缺省值）
    docs/data/scoring/electricity_emission_factors.json  电网因子（Ember 国家级 + 生态环境部中国省级）

口径与决策依据见 docs/CS-DECISIONS.md（D2/D3/D4）。三条不可妥协：

    1. 不编造数值：燃料因子是 IPCC 2006 Vol.2 Table 2.5 的原表数值换算，
       电网因子来自 Ember 年度数据（OWID 镜像）与生态环境部年度公告（curated 录入）。
    2. 单位一律 kgCO2e/kWh，脚本内做量纲断言，超出物理合理区间就整批失败。
    3. 缺就是缺：西藏没有官方省级电网因子（历年公告皆缺），由 §7.4 地理回退
       落到国家级条目，禁止用全国均值冒充省级数据。

用法：

    # OWID/EPA 在境外，需要代理时先设 HTTPS_PROXY 再跑；eGRID 解析需要 openpyxl
    uv run --no-project --with openpyxl python scripts/fetch_emission_factors.py
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
import json
import pathlib
import sys
import time
import urllib.error
import urllib.request
from typing import Any

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# 燃料燃烧因子：IPCC 2006 Guidelines Vol.2 (Energy) Ch.2, Table 2.5
# （Residential and Agriculture/Forestry/Fishing 部门缺省值，kg/TJ，净热值基）
# https://www.ipcc-nggip.iges.or.jp/public/2006gl/vol2.html
#
# CO2e = CO2 + GWP(CH4)·CH4 + GWP(N2O)·N2O
# GWP 用 AR5 100 年值（CH4=28、N2O=265）——现行 UNFCCC 增强透明度框架
# （决定 18/CMA.1 附件）要求各国清单采用的口径。
#
# 2019 Refinement 未改动这些燃料的 CO2 缺省因子，故 5 年刷新周期是安全的
# （见 data-freshness-policy.json 的 rationale）。
# ---------------------------------------------------------------------------
IPCC_RESIDENTIAL = {
    # fuel_key: (CO2 kg/TJ, CH4 kg/TJ, N2O kg/TJ, IPCC 燃料名, confidence)
    "natural_gas": (56100.0, 5.0, 0.1, "Natural Gas", "medium"),
    "lpg": (63100.0, 5.0, 0.1, "Liquefied Petroleum Gases", "medium"),
    "heating_oil": (74100.0, 10.0, 0.6, "Gas/Diesel Oil", "medium"),
    # 居民散煤按 Other Bituminous Coal；居民炉具的 CH4 缺省值就是 300 kg/TJ 这么高，
    # 不同煤种（无烟煤/型煤/兰炭）差异明显，故 confidence 压到 low。
    "solid_fuel": (94600.0, 300.0, 1.5, "Other Bituminous Coal", "low"),
}
GWP100_AR5 = {"CH4": 28.0, "N2O": 265.0}
TJ_PER_KWH = 3.6e-6  # 1 kWh = 3.6 MJ（定义值）
IPCC_URL = "https://www.ipcc-nggip.iges.or.jp/public/2006gl/vol2.html"

# 换算后的物理合理区间（kgCO2e/kWh）。写错单位会差 1–6 个数量级，必然越界。
FUEL_SANITY = {
    "natural_gas": (0.18, 0.23),
    "lpg": (0.20, 0.25),
    "heating_oil": (0.24, 0.29),
    "solid_fuel": (0.30, 0.45),
}

# 燃料因子按国家落条目的国家集合：有居民能源价格覆盖（或即将覆盖）的国家。
# 燃料含碳量是物理性质，逐国重复只为满足 §7.4 的按国解析——不是各国实测值。
# 此表与 scripts/fetch_energy_data.py 的 ISO2_TO_ISO3 保持同步。
PRICE_COVERED_ISO3 = sorted({
    "USA", "CHN",
    "AUT", "BEL", "BGR", "CYP", "CZE", "DEU", "DNK", "EST", "GRC", "ESP", "FIN", "FRA",
    "HRV", "HUN", "IRL", "ITA", "LTU", "LUX", "LVA", "MLT", "NLD", "POL", "PRT", "ROU",
    "SWE", "SVN", "SVK", "ISL", "LIE", "NOR", "CHE", "GBR", "TUR", "MNE", "MKD", "SRB",
    "ALB", "BIH", "MDA", "UKR", "XKX", "GEO",
})

# ---------------------------------------------------------------------------
# 电网因子来源 1：Ember 年度电力数据的 OWID grapher 镜像（CC-BY，无需 key，含 ISO3）。
# Ember 官方下载 URL 每版漂移且 API 需注册，镜像 URL 稳定——见 CS-DECISIONS.md D3。
# 口径为发电碳强度 CO2（不含 CH4/N2O），与全温室气体口径差异一般 <2%，逐条标注。
# ---------------------------------------------------------------------------
OWID_CSV_URL = (
    "https://ourworldindata.org/grapher/carbon-intensity-electricity.csv"
    "?v=1&csvType=full&useColumnShortNames=true"
)

# 电网因子的物理合理区间（kgCO2e/kWh）。下界是 0：接近 100% 水电的小电网
# （如中非共和国）在 Ember 口径下就是 0.0，那是真实数据不是解析错误。
# 解析错误由聚合断言拦：零值国家不得超过 10 个，且 USA/CHN 必须在正常区间。
GRID_SANITY = (0.0, 1.5)
GRID_MAX_ZEROS = 10

# ---------------------------------------------------------------------------
# 电网因子来源 3：EPA eGRID 州级 CO2e 输出排放率（美国官方口径，年度发布）。
# 下载 URL 每个版本都会变——新版本发布后更新这两个常数并重跑
# （版本索引见 https://www.epa.gov/egrid/download-data）。
# ---------------------------------------------------------------------------
EGRID_URL = "https://www.epa.gov/system/files/documents/2025-06/egrid2023_data_rev2.xlsx"
EGRID_SHEET = "ST23"           # 州级表；年份换版时同步改（如 ST24）
EGRID_YEAR = "2023"
LB_PER_KG = 0.45359237         # 定义值

# 与 fetch_energy_data.py 的 US_STATES 保持同步（51 = 50 州 + DC）
US_STATES = [
    "AK", "AL", "AR", "AZ", "CA", "CO", "CT", "DC", "DE", "FL", "GA", "HI", "IA", "ID", "IL",
    "IN", "KS", "KY", "LA", "MA", "MD", "ME", "MI", "MN", "MO", "MS", "MT", "NC", "ND", "NE",
    "NH", "NJ", "NM", "NV", "NY", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT",
    "VA", "VT", "WA", "WI", "WV", "WY",
]

# 生态环境部公告省名（简称）→ 本项目 admin1_code（ISO 3166-2 后缀）。
# 注意 HE=河北、HA=河南、HB=湖北、HN=湖南、NM=内蒙古、SN=陕西——
# 与 Natural Earth postal 互相撞码，见 docs/data/maps/SOURCES.md。
CN_NAME_TO_CODE = {
    "北京": "BJ", "天津": "TJ", "河北": "HE", "山西": "SX", "内蒙古": "NM",
    "辽宁": "LN", "吉林": "JL", "黑龙江": "HL", "上海": "SH", "江苏": "JS",
    "浙江": "ZJ", "安徽": "AH", "福建": "FJ", "江西": "JX", "山东": "SD",
    "河南": "HA", "湖北": "HB", "湖南": "HN", "广东": "GD", "广西": "GX",
    "海南": "HI", "重庆": "CQ", "四川": "SC", "贵州": "GZ", "云南": "YN",
    "西藏": "XZ", "陕西": "SN", "甘肃": "GS", "青海": "QH", "宁夏": "NX",
    "新疆": "XJ",
}


def log(msg: str) -> None:
    print(msg, flush=True)


def http_bytes(url: str, attempts: int = 4) -> bytes:
    last: Exception | None = None
    for i in range(attempts):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "clean-heating-simulator/fetch_emission_factors"}
            )
            with urllib.request.urlopen(req, timeout=120) as r:
                return r.read()
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            last = e
            if i < attempts - 1:
                time.sleep(2 ** i)
    raise RuntimeError("请求失败（重试 %d 次）：%s\n  %s" % (attempts, url, last))


def point(value: float, level: str, code: str, source_name: str, source_url: str,
          retrieved: str, confidence: str, aggregation: str,
          country_iso3: str | None = None, period: str | None = None) -> dict[str, Any]:
    """构造一条 ScoringDataPoint（形状见 docs/src/scoring/dataPoint.ts）。

    排放因子无货币量纲，不带 currency。非 country 级必须带 country_iso3（撞码防护）。
    """
    if level != "country" and not country_iso3:
        raise ValueError("非 country 级条目必须带 country_iso3（%s %s）" % (level, code))
    geography: dict[str, Any] = {"level": level, "code": code}
    if country_iso3 and level != "country":
        geography["country_iso3"] = country_iso3
    pt: dict[str, Any] = {
        "value": round(float(value), 6),
        "geography": geography,
        "source_type": "LOCAL_PUBLIC",
        "source_name": source_name,
        "source_url": source_url,
        "retrieved_at": retrieved,
        "confidence": confidence,
        "aggregation_method": aggregation,
    }
    if period:
        pt["period"] = period
    return pt


def vintage(policy: dict[str, Any], key: str, retrieved: str, source_period: str) -> dict[str, Any]:
    spec = policy[key]
    stale = (_dt.date.fromisoformat(retrieved) + _dt.timedelta(days=spec["refresh_cadence_days"])).isoformat()
    return {
        "retrieved_at": retrieved,
        "source_period": source_period,
        "refresh_cadence_days": spec["refresh_cadence_days"],
        "stale_after": stale,
        "rationale": spec["rationale"],
        "refresh_by": "重跑 " + spec["script"],
    }


# ---------------------------------------------------------------------------
# 燃料燃烧因子
# ---------------------------------------------------------------------------


def build_fuel_factors(retrieved: str, policy: dict[str, Any]) -> dict[str, Any]:
    entries: dict[str, list[dict[str, Any]]] = {}
    for fuel, (co2, ch4, n2o, ipcc_name, conf) in IPCC_RESIDENTIAL.items():
        co2e_per_tj = co2 + ch4 * GWP100_AR5["CH4"] + n2o * GWP100_AR5["N2O"]
        value = co2e_per_tj * TJ_PER_KWH
        lo, hi = FUEL_SANITY[fuel]
        if not (lo <= value <= hi):
            raise RuntimeError("%s 换算出 %.4f kgCO2e/kWh，超出 [%s, %s]——常数或换算写错了" % (fuel, value, lo, hi))
        source_name = (
            "IPCC 2006 GL Vol.2 Ch.2 Table 2.5 (Residential), %s；"
            "CO2e = CO2 + 28·CH4 + 265·N2O（AR5 GWP100，18/CMA.1 口径）" % ipcc_name
        )
        aggregation = (
            "IPCC 缺省值 %.0f/%.0f/%.1f kg{CO2,CH4,N2O}/TJ × 3.6e-6 TJ/kWh。"
            "燃料含碳量是物理性质，各国条目同值，逐国落条目只为满足 §7.4 的按国解析。" % (co2, ch4, n2o)
        )
        entries[fuel] = [
            point(value, "country", iso3, source_name, IPCC_URL, retrieved, conf, aggregation,
                  period="IPCC 2006")
            for iso3 in PRICE_COVERED_ISO3
        ]
        log("  %-12s %.4f kgCO2e/kWh × %d 国" % (fuel, value, len(PRICE_COVERED_ISO3)))

    return {
        "_status": "POPULATED",
        "_owner": "由 scripts/fetch_emission_factors.py 生成，不要手工编辑",
        "_vintage": vintage(policy, "fuel_emission_factors", retrieved,
                            "IPCC 2006 缺省值（2019 Refinement 未改动这些燃料）"),
        "field_key": "fuel_emission_factors",
        "unit": "kgCO2e per kWh（燃料净热值基）",
        "provenance_note": (
            "禁止手挑“典型燃料品质”替代 IPCC 缺省值；散煤因煤种差异 confidence=low。"
            "biomass 不在此文件：净碳强度取决于原料来源与再生周期（carbon_model=biomass_context_dependent），"
            "district_heating/cooling 因子是逐网络的，都没有可引用的公开缺省值——缺就是缺，走 §7.11。"
        ),
        "_gwp_note": "CO2e 采用 AR5 GWP100（CH4=28、N2O=265）。换 GWP 集只需改 GWP100_AR5 常数重跑。",
        "entries": entries,
    }


# ---------------------------------------------------------------------------
# 电网因子
# ---------------------------------------------------------------------------


def fetch_owid_grid(retrieved: str) -> tuple[list[dict[str, Any]], str]:
    raw = http_bytes(OWID_CSV_URL).decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(raw)))
    if not rows:
        raise RuntimeError("OWID CSV 是空的")

    meta_cols = {"Entity", "Code", "Year", "entity", "code", "year"}
    value_cols = [c for c in rows[0].keys() if c not in meta_cols]
    if len(value_cols) != 1:
        raise RuntimeError("OWID CSV 列结构变了：%s（预期 Entity/Code/Year + 1 个数值列）" % list(rows[0].keys()))
    vcol = value_cols[0]

    def col(row: dict[str, str], name: str) -> str:
        return row.get(name) or row.get(name.lower()) or ""

    latest: dict[str, tuple[int, float]] = {}
    for row in rows:
        code = col(row, "Code").strip()
        if len(code) != 3 or not code.isalpha() or code.upper() != code:
            continue  # 跳过 OWID_WRL、区域聚合与空码——规格禁止区域均值替代具体国家
        raw_v = (row.get(vcol) or "").strip()
        if not raw_v:
            continue
        year = int(col(row, "Year"))
        value = float(raw_v)
        if code not in latest or year > latest[code][0]:
            latest[code] = (year, value)

    if len(latest) < 80:
        raise RuntimeError("只解析出 %d 个国家，OWID CSV 结构可能变了" % len(latest))
    usa_g = latest.get("USA", (0, -1.0))[1]
    if not (150.0 <= usa_g <= 700.0):
        raise RuntimeError("USA = %s gCO2/kWh，不在合理区间——单位或列解析错了" % usa_g)

    zeros = [iso3 for iso3, (_, g) in latest.items() if g / 1000.0 < 0.005]
    if len(zeros) > GRID_MAX_ZEROS:
        raise RuntimeError("有 %d 个国家电网因子近零（%s…），像是解析错误" % (len(zeros), ", ".join(sorted(zeros)[:5])))
    chn_g = latest.get("CHN", (0, -1.0))[1]
    if not (300.0 <= chn_g <= 800.0):
        raise RuntimeError("CHN = %s gCO2/kWh，不在合理区间——单位或列解析错了" % chn_g)

    points: list[dict[str, Any]] = []
    years: set[int] = set()
    for iso3 in sorted(latest):
        year, g_per_kwh = latest[iso3]
        kg = g_per_kwh / 1000.0
        if not (GRID_SANITY[0] <= kg <= GRID_SANITY[1]):
            raise RuntimeError("grid %s = %.4f kgCO2e/kWh 越界" % (iso3, kg))
        years.add(year)
        points.append(point(
            kg, "country", iso3,
            "Ember Yearly Electricity Data（经 Our World in Data carbon-intensity-electricity 镜像），%d" % year,
            OWID_CSV_URL, retrieved, "high",
            "Ember 发电碳强度年度值；口径为 CO2（不含 CH4/N2O），与全温室气体口径差异一般 <2%",
            period=str(year),
        ))
    return points, "%d–%d" % (min(years), max(years))


def load_cn_grid(repo: pathlib.Path, retrieved: str) -> tuple[list[dict[str, Any]], str]:
    curated_path = repo / "scripts" / "curated_cn_grid_ef.json"
    if not curated_path.exists():
        log("  未找到 %s，跳过中国省级电网因子（国家级 Ember CHN 仍在）" % curated_path.name)
        return [], ""
    curated = json.loads(curated_path.read_text(encoding="utf-8"))

    # 与地图数据交叉校验省名→码映射：简称必须是 geojson 全称的前缀且码一致。
    geo = json.loads((repo / "docs/data/maps/admin1-cn-us.geojson").read_text(encoding="utf-8"))
    full_by_code = {
        f["properties"]["admin1_code"]: f["properties"]["name_zh"]
        for f in geo["features"] if f["properties"]["country_iso3"] == "CHN"
    }
    for short, code in CN_NAME_TO_CODE.items():
        full = full_by_code.get(code)
        if not full or not full.startswith(short):
            raise RuntimeError("省名映射自检失败：%s→%s 与地图 name_zh=%r 对不上" % (short, code, full))

    seen: set[str] = set()
    points: list[dict[str, Any]] = []
    for row in curated["provinces"]:
        name, value = row["name_zh"], float(row["value"])
        code = CN_NAME_TO_CODE.get(name)
        if not code:
            raise RuntimeError("curated 里出现未知省名：%r" % name)
        if code in seen:
            raise RuntimeError("curated 里省份重复：%s" % name)
        seen.add(code)
        if not (0.05 <= value <= 1.2):
            raise RuntimeError("%s = %s kgCO2/kWh 越界——curated 数据有误" % (name, value))
        points.append(point(
            value, "admin1", code,
            "%s，%s" % (curated["announcement_title"], curated["scope"]),
            curated["announcement_url"], curated.get("retrieved_at", retrieved), "high",
            "生态环境部/国家统计局年度公告省级平均因子；口径为 CO2。curated 录入（PDF 无法程序化抓取），"
            "校验记录见 scripts/curated_cn_grid_ef.json 的 _curation_note",
            country_iso3="CHN", period=str(curated["factor_year"]),
        ))
    if len(points) < 30:
        raise RuntimeError("中国省级因子只有 %d 条，官方表至少 30 省" % len(points))
    missing = sorted(set(CN_NAME_TO_CODE.values()) - seen)
    if missing:
        log("  官方无数据的省（由 §7.4 回退到国家级，不代填）：%s" % ", ".join(missing))
    points.sort(key=lambda p: p["geography"]["code"])
    return points, str(curated["factor_year"])


def fetch_egrid_states(retrieved: str) -> list[dict[str, Any]]:
    try:
        import openpyxl  # noqa: PLC0415 —— 只有 eGRID 需要它；缺了就明确报错
    except ImportError:
        raise RuntimeError("解析 eGRID 需要 openpyxl：uv run --no-project --with openpyxl python %s" % __file__)
    import io as _io

    raw = http_bytes(EGRID_URL)
    wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
    if EGRID_SHEET not in wb.sheetnames:
        raise RuntimeError("eGRID 文件里没有 %s 表——版本换了，更新 EGRID_URL/EGRID_SHEET/EGRID_YEAR" % EGRID_SHEET)
    ws = wb[EGRID_SHEET]
    rows = ws.iter_rows(values_only=True)
    next(rows)                      # 第 1 行是全称表头
    codes = [str(c or "") for c in next(rows)]   # 第 2 行是字段码
    try:
        i_state = codes.index("PSTATABB")
        i_rate = codes.index("STC2ERTA")   # State annual CO2e total output emission rate (lb/MWh)
    except ValueError:
        raise RuntimeError("ST 表字段码变了（找不到 PSTATABB/STC2ERTA）：%s" % codes[:30])

    by_state: dict[str, float] = {}
    for row in rows:
        st = str(row[i_state] or "").strip()
        if st in US_STATES and row[i_rate] is not None:
            # data_only=True 下单元格是数值；经 str 桥接是为了让非数值内容显式炸掉而不是被静默跳过
            by_state[st] = float(str(row[i_rate])) * LB_PER_KG / 1000.0   # lb/MWh → kg/kWh

    missing = sorted(set(US_STATES) - set(by_state))
    if missing:
        raise RuntimeError("eGRID 缺州：%s" % ", ".join(missing))
    # 锚点断言：西弗吉尼亚煤电重、华盛顿水电重，写错换算或取错列必然越界
    if not (0.6 <= by_state["WV"] <= 1.2):
        raise RuntimeError("WV = %.3f kg/kWh，锚点越界" % by_state["WV"])
    if not (0.005 <= by_state["WA"] <= 0.2):
        raise RuntimeError("WA = %.3f kg/kWh，锚点越界" % by_state["WA"])

    points = [
        point(
            v, "admin1", st,
            "EPA eGRID%s（rev2）ST 表 STC2ERTA：州年度 CO2e 总输出排放率" % EGRID_YEAR,
            EGRID_URL, retrieved, "high",
            "lb/MWh × %.8f ÷ 1000 → kgCO2e/kWh；口径为 CO2e（含 CH4/N2O）" % LB_PER_KG,
            country_iso3="USA", period=EGRID_YEAR,
        )
        for st, v in sorted(by_state.items())
    ]
    return points


def build_electricity_factors(repo: pathlib.Path, retrieved: str, policy: dict[str, Any]) -> dict[str, Any]:
    log("  Ember（OWID 镜像）国家级……")
    country_points, ember_period = fetch_owid_grid(retrieved)
    log("  %d 个国家（%s）" % (len(country_points), ember_period))
    cn_points, mee_year = load_cn_grid(repo, retrieved)
    if cn_points:
        log("  中国省级 %d 条（%s 年因子）" % (len(cn_points), mee_year))
    us_points = fetch_egrid_states(retrieved)
    log("  美国州级 %d 条（eGRID%s）" % (len(us_points), EGRID_YEAR))

    source_period = ember_period + (("/" + mee_year + " (MEE)") if mee_year else "") + "/%s (eGRID)" % EGRID_YEAR
    return {
        "_status": "POPULATED",
        "_owner": "由 scripts/fetch_emission_factors.py 生成，不要手工编辑",
        "_vintage": vintage(policy, "electricity_emission_factors", retrieved, source_period),
        "field_key": "electricity_emission_factors",
        "unit": "kgCO2e per kWh",
        "provenance_note": (
            "禁止把中国「电力二氧化碳排放因子」与「电力碳足迹因子」混用——后者是生命周期口径，数值系统性偏高。"
        ),
        "_scope_note": (
            "气体覆盖口径不一且逐条已标注：中国省级（MEE）与国家级（Ember）为 CO2，"
            "美国州级（eGRID）为 CO2e（含 CH4/N2O），差异一般 <2%。"
            "西藏无官方省级因子，§7.4 回退到国家级 CHN 条目。"
        ),
        "entries": {"grid": cn_points + us_points + country_points},
    }


# ---------------------------------------------------------------------------


def main() -> int:
    repo = pathlib.Path(__file__).resolve().parent.parent
    out_dir = repo / "docs" / "data" / "scoring"
    retrieved = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d")
    policy = json.loads((repo / "docs/data/data-freshness-policy.json").read_text(encoding="utf-8"))["datasets"]

    log("[1/2] 燃料燃烧因子（IPCC 2006 缺省值，本地换算）")
    fuel = build_fuel_factors(retrieved, policy)

    log("[2/2] 电网因子")
    elec = build_electricity_factors(repo, retrieved, policy)

    for name, payload in (("fuel_emission_factors.json", fuel), ("electricity_emission_factors.json", elec)):
        p = out_dir / name
        p.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        total = sum(len(v) for v in payload["entries"].values())
        log("  wrote %s  %s bytes（共 %d 条）" % (name, format(p.stat().st_size, ","), total))
    return 0


if __name__ == "__main__":
    sys.exit(main())
